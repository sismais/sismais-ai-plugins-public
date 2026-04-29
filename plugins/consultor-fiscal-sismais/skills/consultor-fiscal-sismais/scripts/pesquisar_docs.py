#!/usr/bin/env python3
"""
Pesquisa textual na base de documentos fiscais (doc_sources/).
Suporta: PDF, CSV, JSON, Excel (.xlsx/.xls), TXT, MD, XML.

Uso:
    python pesquisar_docs.py "termo de busca"
    python pesquisar_docs.py "CFOP 5102" --pasta notas_tecnicas
    python pesquisar_docs.py "IBS CBS" --tipo pdf
    python pesquisar_docs.py "regra validação" --max 20
"""

import argparse
import csv
import importlib
import json
import os
import re
import subprocess
import sys
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parent.parent
DOC_SOURCES = SKILL_DIR / "doc_sources"

REQUIRED_PACKAGES = {
    "pdfplumber": "pdfplumber",
    "openpyxl": "openpyxl",
}


def _ensure_dependencies():
    """Instala pacotes ausentes automaticamente. Avisa o usuário se falhar."""
    missing = []
    for import_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            importlib.import_module(import_name)
        except ImportError:
            missing.append((import_name, pip_name))
    if not missing:
        return
    nomes = [pip for _, pip in missing]
    print(f"[INFO] Instalando dependências: {', '.join(nomes)}...", file=sys.stderr)
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", *nomes]
        )
    except (subprocess.CalledProcessError, FileNotFoundError, OSError) as e:
        print(
            f"\n[ERRO] Não foi possível instalar automaticamente: {', '.join(nomes)}\n"
            f"       Motivo: {e}\n"
            f"       Instale manualmente com:\n\n"
            f"         pip install {' '.join(nomes)}\n",
            file=sys.stderr,
        )
        sys.exit(1)
    failed = []
    for import_name, pip_name in missing:
        try:
            importlib.import_module(import_name)
        except ImportError:
            failed.append(pip_name)
    if failed:
        print(
            f"\n[ERRO] Pacotes instalados mas não importáveis: {', '.join(failed)}\n"
            f"       Verifique se o Python em uso é o correto:\n"
            f"         {sys.executable}\n"
            f"       Tente instalar manualmente:\n\n"
            f"         \"{sys.executable}\" -m pip install {' '.join(failed)}\n",
            file=sys.stderr,
        )
        sys.exit(1)


_ensure_dependencies()

SUPPORTED_EXTENSIONS = {
    "pdf": [".pdf"],
    "csv": [".csv"],
    "json": [".json"],
    "excel": [".xlsx", ".xls"],
    "texto": [".txt", ".md", ".xml", ".xsd"],
}
ALL_EXTENSIONS = [ext for exts in SUPPORTED_EXTENSIONS.values() for ext in exts]


def extrair_texto_pdf(filepath: Path) -> str:
    import pdfplumber
    with pdfplumber.open(filepath) as pdf:
        return "\n".join(
            page.extract_text() or "" for page in pdf.pages
        )


def extrair_texto_excel(filepath: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    linhas = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            linhas.append(" | ".join(str(c) if c is not None else "" for c in row))
    wb.close()
    return "\n".join(linhas)


def extrair_texto_csv(filepath: Path) -> str:
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc) as f:
                reader = csv.reader(f)
                return "\n".join(" | ".join(row) for row in reader)
        except (UnicodeDecodeError, csv.Error):
            continue
    return "[ERRO] Não foi possível ler o CSV"


def extrair_texto_json(filepath: Path) -> str:
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"[ERRO] {e}"


def extrair_texto_plaintext(filepath: Path) -> str:
    encodings = ["utf-8", "latin-1", "cp1252"]
    for enc in encodings:
        try:
            return filepath.read_text(encoding=enc)
        except UnicodeDecodeError:
            continue
    return "[ERRO] Não foi possível ler o arquivo"


def extrair_texto(filepath: Path) -> str:
    ext = filepath.suffix.lower()
    if ext == ".pdf":
        return extrair_texto_pdf(filepath)
    elif ext in (".xlsx", ".xls"):
        return extrair_texto_excel(filepath)
    elif ext == ".csv":
        return extrair_texto_csv(filepath)
    elif ext == ".json":
        return extrair_texto_json(filepath)
    else:
        return extrair_texto_plaintext(filepath)


def buscar_arquivos(pasta: str = None, tipo: str = None) -> list[Path]:
    base = DOC_SOURCES
    if pasta:
        base = DOC_SOURCES / pasta
        if not base.exists():
            print(f"[AVISO] Pasta não encontrada: {base}", file=sys.stderr)
            return []

    extensoes = ALL_EXTENSIONS
    if tipo and tipo in SUPPORTED_EXTENSIONS:
        extensoes = SUPPORTED_EXTENSIONS[tipo]

    arquivos = []
    for ext in extensoes:
        arquivos.extend(base.rglob(f"*{ext}"))
    return sorted(arquivos)


def pesquisar(termo: str, pasta: str = None, tipo: str = None, max_resultados: int = 10, contexto: int = 2):
    pattern = re.compile(re.escape(termo), re.IGNORECASE)
    arquivos = buscar_arquivos(pasta, tipo)

    if not arquivos:
        print(f"Nenhum arquivo encontrado em {DOC_SOURCES / (pasta or '')}")
        return

    total_matches = 0
    print(f"Pesquisando '{termo}' em {len(arquivos)} arquivo(s)...\n")

    for filepath in arquivos:
        texto = extrair_texto(filepath)
        if texto.startswith("[ERRO]"):
            print(f"  {filepath.relative_to(SKILL_DIR)}: {texto}")
            continue

        linhas = texto.split("\n")
        matches = []
        for i, linha in enumerate(linhas):
            if pattern.search(linha):
                inicio = max(0, i - contexto)
                fim = min(len(linhas), i + contexto + 1)
                matches.append((i + 1, linhas[inicio:fim]))

        if matches:
            rel_path = filepath.relative_to(SKILL_DIR)
            print(f"{'='*80}")
            print(f"ARQUIVO: {rel_path}")
            print(f"MATCHES: {len(matches)}")
            print(f"{'='*80}")

            for line_num, ctx_lines in matches[:max_resultados]:
                print(f"\n  Linha ~{line_num}:")
                for cl in ctx_lines:
                    marker = ">>>" if pattern.search(cl) else "   "
                    print(f"  {marker} {cl.strip()}")

            total_matches += len(matches)

            if total_matches >= max_resultados:
                print(f"\n... limite de {max_resultados} resultados atingido.")
                break

    print(f"\nTotal: {total_matches} ocorrência(s) encontrada(s).")


def listar_base():
    print(f"Base de documentos: {DOC_SOURCES}\n")
    if not DOC_SOURCES.exists():
        print("  [VAZIA] Pasta doc_sources/ não encontrada.")
        return

    for pasta in sorted(DOC_SOURCES.iterdir()):
        if pasta.is_dir():
            arquivos = list(pasta.rglob("*"))
            arquivos = [a for a in arquivos if a.is_file()]
            print(f"  {pasta.name}/ ({len(arquivos)} arquivo(s))")
            for arq in sorted(arquivos)[:10]:
                print(f"    - {arq.name}")
            if len(arquivos) > 10:
                print(f"    ... e mais {len(arquivos) - 10}")


def main():
    parser = argparse.ArgumentParser(
        description="Pesquisa textual na base de documentos fiscais"
    )
    parser.add_argument("termo", nargs="?", help="Termo de busca")
    parser.add_argument("--pasta", help="Subpasta em doc_sources/ (ex: notas_tecnicas, tabelas)")
    parser.add_argument("--tipo", choices=SUPPORTED_EXTENSIONS.keys(), help="Filtrar por tipo de arquivo")
    parser.add_argument("--max", type=int, default=10, help="Máximo de resultados (default: 10)")
    parser.add_argument("--contexto", type=int, default=2, help="Linhas de contexto (default: 2)")
    parser.add_argument("--listar", action="store_true", help="Listar arquivos na base")

    args = parser.parse_args()

    if args.listar:
        listar_base()
        return

    if not args.termo:
        parser.print_help()
        return

    pesquisar(args.termo, args.pasta, args.tipo, args.max, args.contexto)


if __name__ == "__main__":
    main()
